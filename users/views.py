from sre_constants import SUCCESS
from django.shortcuts import render, redirect
from django.contrib import messages
from socialmedia.settings import LOGIN_REDIRECT_URL
from .models import Profile
from posts.models import Post
from django.views import View
from django.views import generic
from django.urls import reverse_lazy
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from django.http import HttpResponse
import datetime
from django.contrib.auth.forms import UserChangeForm
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from django.views.generic import UpdateView
from .forms import UserRegisterForm, UserUpdateForm, ProfileUpdateForm

def update_user_data(user):
    Profile.objects.update_or_create(user=user, defaults = {'mob': user.profile})

def register(request):
    form_class = UserRegisterForm
    profile_form_class = ProfileUpdateForm
    form = form_class(request.POST or None)
    profile_form = profile_form_class(request.POST or None)
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        profile_form = ProfileUpdateForm(request.POST)
        if form.is_valid() and profile_form.is_valid():
            user = form.save()
            profile = profile_form.save(commit = False)
            if profile.user_id  is None :
                profile.user_id = user.id
            profile.save()
            user.refresh_from_db()
            user.profile.mob = form.cleaned_data.get('mob')
            update_user_data(user)
            user.save()
            messages.success(request, f'Account created for {user.username} !')
            return redirect('login')
            
        else:
            form = UserRegisterForm()
            profile_form = ProfileUpdateForm()
    context = { 'form': form, 'profile_form': profile_form }
    return render(request, 'users/register.html', context)

# def profileupdate (request):
#     if request.method == 'POST':
#         u_form = UserUpdateForm(request.POST, instance = request.user)
#         p_form = ProfileUpdateForm( request.POST,
# 								 	request.FILES,
# 									instance = request.user.profile)
#         if u_form.is_valid() and p_form.is_valid():
#             u_form.save()
#             p_form.save()
#             messages.success(request, f'Your account has been updated !')
#             return redirect('profile')
#         else:
#             print(u_form.errors)
#     else:
#         u_form = UserUpdateForm(instance = request.user)
#         p_form = ProfileUpdateForm(instance = request.user.profile)
		
#     context = {
# 		'u_form': u_form,
# 		'p_form': p_form,
# 	}
#     return render(request, 'users/profile.html', context)

# class UserEditView(generic.UpdateView):
#     form_class = UserChangeForm
#     template_name = 'users/edit_profile.html'
#     success_url = reverse_lazy('home')

#     def get_object(self):
#         return self.request.user

# Check why not working
class AddFollower(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        profile = Profile.objects.get(pk=pk)
        profile.followers.add(request.user)

        return redirect('profile', pk=profile.pk)

class RemoveFollower(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        profile = Profile.objects.get(pk=pk)
        profile.followers.remove(request.user)

        return redirect('profile', pk=profile.pk)

class ProfileView (View):
    def get(self, request, pk, *args, **kwargs):
        profile = Profile.objects.get(pk= pk)
        user = profile.user
        own_posts = Post.objects.filter (op_name = user).order_by('-pub_date')
        followers = profile.followers.all()
        no_of_followers = len (followers)

        if len (followers) == 0 :
            is_following = False
        #Checking if is followig
        for follower in followers :
            if follower == request.user :
                is_following = True 
                break
            else :
                is_following = False

        context = {
            'user' : user,
            'profile': profile,
            'own_posts': own_posts,
            'no_of_followers' : no_of_followers,
            'is_following' : is_following
        }

        return render (request, 'users/profile.html', context)

    # def post (self, request, pk, *args, **kwargs) :
    #     return render (request, 'users/profile.html')

# Try making it a profile def 
# class ProfileEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
#     model = Profile
#     # fields = ['name', 'location', 'profile_pic', 'mobile']
#     # template_name = 'users/profile_edit.html'

#     def get_success_url(self):
#         pk = self.kwargs['pk']
#         return reverse_lazy('profile', kwargs={'pk': pk})

#     def test_func(self):
#         profile = self.get_object()
#         return self.request.user == profile
    
#     def post (self, request, *args, **kwargs):
#         u_form = UserUpdateForm(instance = request.user)
#         p_form = ProfileUpdateForm(instance = request.user.profile)
#         context = {
#             'u_form': u_form,
#             'p_form': p_form
#         }
#         print('get request works')
#         return render (request, 'users/profile_edit.html', context)

class EditProfileView(generic.UpdateView):
    model = Profile
    # template_name = 'users/edit_profile.html'
    form_class = ProfileUpdateForm
    # fields = ['profile_pic', 'mobile', 'location', 'bio']
    def post (self, request, *args, **kwargs) :
        form = ProfileUpdateForm(instance = request.user.profile)
        if form.is_valid() :
            form.save()
        else :
            print (form.errors)
        context = {
            'form': form
        }
        # success_url = 'home'
        return render (request, 'users/edit_profile.html', context )

    def get_success_url(self):
        pk = self.kwargs['pk']
        return reverse_lazy('profile', kwargs={'pk': pk})

    def test_func(self):
        profile = self.get_object()
        return self.request.user == profile

class UserSearch(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        query = self.request.GET.get('query')
        profile_list = Profile.objects.filter(
            Q(user__username__icontains=query)
        )
        context = {
            'profile_list': profile_list,
        }
        return render(request, 'users/user_search.html', context)

@login_required 
def searchpage (request):
    #To find most popular users
    profiles = Profile.objects.order_by('no_of_followers')[:5]
    user = request.user
    context = {
        'profiles': profiles,
        'user': user
    }
    return render (request, 'users/searchpage.html', context)

class ListFollowers(View):
    def get(self, request, pk, *args, **kwargs):
        profile = Profile.objects.get(pk=pk)
        followers = profile.followers.all()
        no_of_followers = len(followers)
        context = {
            'profile': profile,
            'followers': followers,
            'no_of_followers': no_of_followers
        }
        return render(request, 'users/followers_list.html', context)

def exportToExcel (request): 
    # wb = load_workbook('Profiles.xlsx')
    wb = Workbook(write_only = True)
    sheet = wb.create_sheet()
    columnheadings = ['Username','email', 'Mobile', 'Location', 'No of followers', 'Bio']
    # for i in range (1, 5):
    #     sheet.cell(row = 1, column = i).value = columnheadings[i-1]
    sheet.append(columnheadings)
    profiles = Profile.objects.all()
    for profile in profiles.iterator() :
        username = profile.user.username
        email = profile.user.email
        mobile = profile.mobile
        location = profile.location
        no_of_followers = profile.no_of_followers
        bio = profile.bio
        data = [username, email,  mobile, location, no_of_followers, bio]
        sheet.append(data)
    wb.save('Profiles.xlsx')
    return redirect ('searchpage')

#Make a mutuals view (ie if person in profile.follwers and request.user.followers then  mutaul +=1 where 
# profile = Profile.objects.filter(pk = pk))
